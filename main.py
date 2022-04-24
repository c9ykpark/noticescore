# 사전에 설치되어 있어야 하는 라이브러리
# pip3 install openpyxl

from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# 0. 변수 설정
score_student_id_col_num = 3
mailing_list_student_id_col_num = 0
mailing_list_student_email_col_num = 2


def find_row_num(score, student_id):
    # 전체 성적에서 student_id에 해당하는 행을 가져온다.
    # 있으면 행 번호 리턴, 없으면 -1 리턴
    found = 0

    for row in score:
        if row[score_student_id_col_num] == student_id:
            return found
        found = found + 1

    return -1


def create_mail_text(score):
    text = """
        안녕하세요.<br/>
    
        중간고사 시험점수를 안내드립니다.<br/><br/>
        <table border='1' cellpadding='5' style='border-collapse:collapse;'>
        <tr align='center'>
            <th rowspan='2'>학번</th>
            <th rowspan='2'>이름</th>
            <th colspan='3'>이론</th>
            <th colspan='4'>파이썬</th>
            <th rowspan='2'>합계<br/>(100점)</th>
        </tr>
            <tr align='center'>
            <th>1번<br/>(20점)</th>
            <th>2번<br/>(30점)</th>
            <th>3번<br/>(18점)</th>
            <th>1번<br/>(6점)</th>
            <th>2번<br/>(6점)</th>
            <th>3번<br/>(14점)</th>
            <th>4번<br/>(6점)</th>
        </tr>
        <tr align='center'>
            <td>{0}</td>
            <td>{1}</td>
            <td>{2}</td>
            <td>{3}</td>
            <td>{4}</td>
            <td>{5}</td>
            <td>{6}</td>
            <td>{7}</td>
            <td>{8}</td>
            <td>{9}</td>
        </tr>
        </table>
        <p> 
        채점결과에 대한 문의사항은 실습시간에 교수님께 문의 바랍니다.<br/><br/>
        남은 수업 잘 마무리하여 좋은 결과가 있길 바랍니다.
        """
    text = text.format(score[3], score[4], # 학번, 이름
                score[5], score[6], score[7], # 이론 문제
                score[8], score[9], score[10], score[11], # 파이썬 문제
                score[12]) # 합계
    print(text)

    return text


if __name__ == '__main__':
    # 2. 성적 전체를 받아와 total_score 리스트에 저장
    score_wb = load_workbook('score.xlsx', data_only=True)
    score_ws = score_wb['Sheet1']
    total_score = []
    for row in score_ws.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
        total_score.append(row_value)
    score_wb.close()

    # 3-1. 메일링 리스트에서 한 row 씩 읽으면서
    mailing_list_wb = load_workbook('mail-list.xlsx', data_only=True)
    mailing_list_ws = mailing_list_wb['Sheet1']
    for student_row in mailing_list_ws.rows:
        student_value = []
        for student_cell in student_row:
            student_value.append(student_cell.value)

        if student_value[mailing_list_student_email_col_num] == None:
            # 메일링 리스트에 이메일 정보가 저장되어 있지 않을 경우 바로 pass
            continue

        # 3-2. total_score 에서 학번이 일치하는 row 가 있으면
        row_num = find_row_num(total_score, str(student_value[mailing_list_student_id_col_num]))
        if row_num == -1:
            print("[-] Not found", student_value[mailing_list_student_id_col_num])
            continue
        else:
            # 4. 작성한 내용으로 메일을 보냄
            email_from = "(Sender's E-mail)"
            email_to = student_value[mailing_list_student_email_col_num]
            email_subject = "(Title)"

            email_content = MIMEMultipart("alternative")
            email_content.set_charset('UTF-8')
            email_content['From'] = email_from
            email_content['To'] = email_to
            email_content['Subject'] = email_subject

            # 3-3. 해당 정보를 바탕으로 메일 텍스트를 작성함
            text = create_mail_text(total_score[row_num])
            text_part = MIMEText(text, 'html', 'UTF-8')
            email_content.attach(text_part)

            smtp = smtplib.SMTP("smtp.gmail.com", 587)
            smtp.set_debuglevel(True)
            smtp.starttls()
            smtp.login("(Sender's E-mail)", "(App-Password)")
            smtp.sendmail(email_from, email_to, email_content.as_string())

            smtp.quit()
            print("[+] Send", student_value[mailing_list_student_id_col_num],
                  " to ", student_value[mailing_list_student_email_col_num])

    mailing_list_wb.close()

